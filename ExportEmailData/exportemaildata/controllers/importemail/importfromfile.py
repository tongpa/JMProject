#!/usr/bin/python

import threading
import time
import openpyxl
import os
import os.path

from tg.configuration import AppConfig, config
from exportemaildata import model
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.sql import update
from exportemaildata.controllers.utility import Utility; 

import sys
import logging;
log = logging.getLogger(__name__);

__all__ = ['importDataThread']
class importDataThread(threading.Thread):
    
    
    utility = Utility();
    
    def __init__(self, threadID,   pathFile,importEmail = None):
        reload(sys).setdefaultencoding('utf8')
        threading.Thread.__init__(self)
        
        log.info("init import Data thread");
        self.threadID = threadID
        self.importEmail = None;
        self.idExport = None;
        if(importEmail == None):        
            #self.model = model
            self.pathFile = pathFile;
        else:
            print importEmail.id_export_email;
            self.idExport = importEmail.id_export_email;
            self.importEmail = importEmail;
            self.pathFile = importEmail.path_file;
        
        self.some_engine = create_engine(config['sqlalchemy.first.url'] );
        # create a configured "Session" class
        self.Session = sessionmaker(bind=self.some_engine)
        # create a Session
        self.session = self.Session();
        
        
        
    def run(self):
        
        log.info("Starting importData : " + str(self.threadID));
        
        import datetime

        start = datetime.datetime.now()
        print start 
        
        #step : 1 importData
       
        exportEmail = self.importData(self.threadID ,self.pathFile); 
        
        finish_export = datetime.datetime.now()
        print start;
        print finish_export
       
        #step : 2 checkData Same Old Email
        #exportEmail = model.ExportEmail.getId(5);
        log.info("start check email Duplicate from database");
        self.checkEmailDuplicate(exportEmail);
        
        
        finish_import = datetime.datetime.now() 
        print start;
        print finish_export
        print finish_import
        log.info("Finish importData : " + str(self.threadID));
     
        
    def importData(self,threadName,   pathFile):
        reload(sys).setdefaultencoding('utf8')
        
        log.info("read file " + str(pathFile).encode());
         
        workbook = openpyxl.load_workbook(filename = pathFile, use_iterators = True)
        #worksheets = workbook.get_sheet_names()
        #worksheet = workbook.get_sheet_by_name('Sheet13')    
        worksheet =  workbook.active
        log.info( str(worksheet.calculate_dimension()) );
        
        self.email = {};
        self.pid = {};
        self.email_empty=[];
        self.mobile_empty = [];
        self.same_email = [];
        self.same_pid = [];
        self.used_email = [];
        
        
        path = r'C:\temp\demo1.xlsx';
        
        # an Engine, which the Session will use for connection
        # resources
        
        
        # work with sess
        if(self.importEmail == None):
            exportEmail = model.ExportEmail();
            
            exportEmail.file_name = "data1";
            exportEmail.path_file = "ddddd.exls";
            exportEmail.error_path_file = path;
    
            exportEmail.total_row = 0;
            exportEmail.insert_row = 0;
            exportEmail.error_row = 0;
            
            exportEmail.same_old_row = 0;
            exportEmail.insert_real_row = 0;
            
            self.session.add(exportEmail);
            self.session.flush() ;
            self.session.commit();
        else:
            print  self.idExport 
            
            exportEmail = self.session.query(model.ExportEmail).get(self.idExport);
            path = exportEmail.error_path_file;
            log.info( 'get export data id ' + str(exportEmail.id_export_email) );
            log.info( 'error path file' + str(path) );
        
        total = 0;
        exportEmail.id_status_export = 3 ;
        log.info( 'start insert table email data');
        for row in worksheet.iter_rows():
            total = total +1;
            emaildata = model.EmailData();
            emaildata.prefix =  row[0].value;
            emaildata.id_export_email = exportEmail.id_export_email;
            emaildata.firstname_thai =  row[1].value;
            emaildata.lastname_thai =  row[2].value;
            emaildata.firstname_eng =  row[3].value;
            emaildata.lastname_eng =  row[4].value;
           
            emaildata.sex =  row[5].value;
            emaildata.birthdate =  row[6].value;
    
            emaildata.pid =  row[7].value;
            emaildata.passport =  row[8].value;
            emaildata.countryname =  row[9].value;
            emaildata.house_no =  row[10].value;
            emaildata.building_village =  row[11].value;
            emaildata.moo =  row[12].value;
            emaildata.soi =  row[13].value;
            emaildata.road =  row[14].value;
            emaildata.county =  row[15].value;
            emaildata.city =   row[16].value;
            emaildata.province =   row[17].value;
            emaildata.postcode=  row[18].value;
            emaildata.mobile=  row[19].value;
            emaildata.telephone=   row[20].value;
            
            if(row[21].value and self.utility.checkEmail(row[21].value)):
                emaildata.email=   self.utility.lastReplace(row[21].value.lower(),'.')    ;
            else:
                emaildata.email = None;
                log.info("row : " + str(total) + "  : " +str(row[21].value) );
            
            emaildata.housing_type=   row[22].value;
            emaildata.category=   row[23].value;
            emaildata.salary=  row[24].value;
            emaildata.education=   row[25].value;
            
            #check email is not empty
            if(emaildata.email is not None and str(emaildata.email).strip() != ''  ):
                #check email is same
                if (self.email.get(emaildata.email) is None) :
                    self.email[emaildata.email] = emaildata;
                     
                else:
                    self.same_email.append(emaildata);
            else:
                emaildata.email=   row[21].value    ;
                self.email_empty.append(emaildata);
                
            
            if(total % 500 ) == 0 :
                log.info( total );   
            
            
        log.info("success check data : ");         
        
        log.info( "same email before " + str( len(self.same_email)));
        log.info( "used email before " + str( len(self.email)) );
        #remove same email
        for email in self.same_email:
            sameemail = self.email.get(email.email);
            log.info( "same email : " + str(email.email));
            if(sameemail):
                del self.email[email.email];
                self.same_email.append(sameemail);     
             
        
        log.info("check mobile is empty");
        log.info("leng email : " + str(len(self.email)));
        for key_email in self.email.keys():
            email_value = self.email.get(key_email);
            
            if(email_value.mobile is None or str(email_value.mobile).strip() == '' ):
                del self.email[key_email];
                self.mobile_empty.append(email_value);     
        
        
        row =1000;
        count = 0;
        log.info( "start insert");
        exportEmail.id_status_export = 4 ;
        for key_email in self.email.keys():
            useEmail = self.email[key_email];
            
            try:
                
                self.session.add(useEmail);
                self.session.flush() ;
                
                if(count % row == 0):
                    log.info(".....commit");
                    #log.info("commit at : " + str(count));
                    self.session.commit();
                    #print "commit ";
                    
            except Exception as e:
                 
                log.info( "---error---");
                log.info(useEmail);
                #log.info( "email :" + str(useEmail.email ));
                print e;
            
            count = count + 1;
            
            
           
        exportEmail.total_row = total;
        exportEmail.insert_row = count;
        exportEmail.error_row = total- count;
        
        self.session.commit();
        
        log.info( "used email " + str( len(self.email)));
        log.info( "same email " + str( len(self.same_email)));
        log.info( "same ipd   " + str( len(self.same_pid)));
        
        log.info("---------------------");
        log.info("------same_email---------------" + str(len(self.same_email)));
        log.info("------same_pid---------------" + str(len(self.same_pid)));
        log.info("------email_empty---------------" + str(len(self.email_empty)));        
        log.info("------mobile_empty---------------" + str(len(self.mobile_empty)));
        
        workbook = openpyxl.Workbook();
        
        sheet = workbook.get_active_sheet();
        workbook.remove_sheet(sheet);
        
        if(len(self.same_email) >0):
            self.writeToExcel(workbook,self.same_email,path,'same email');
        
        if(len(self.same_pid) >0):
            self.writeToExcel(workbook,self.same_pid,path,'same pid');
        
        if(len(self.email_empty) > 0):
            self.writeToExcel(workbook,self.email_empty,path,'email empty');
            
        if(len(self.mobile_empty) > 0):
            self.writeToExcel(workbook,self.mobile_empty,path,'mobile empty');
        
        exportEmail.id_status_export = 5 ;    
        return exportEmail;
            
        #workbook.save(path);
    
    def writeToExcel(self,workbook,data,fileName,sheedName):
        reload(sys).setdefaultencoding('utf8')
        #wb = openpyxl.Workbook();
        
        wb = workbook;
        
        #sheet = wb.get_active_sheet()
        sheet = wb.create_sheet()
        sheet.title = sheedName;
        
        log.info("add sheed : " + str(sheedName));
        
        i =1;
        for v in data:
            cell_C1 = sheet.cell( row=i, column=1 );
            cell_C1.value = v.prefix;
            
            cell_C2 = sheet.cell( row=i, column=2 );
            cell_C2.value = v.firstname_thai;
    
            cell_C3 = sheet.cell( row=i, column=3 );
            cell_C3.value = v.lastname_thai;
            
            cell_C4 = sheet.cell( row=i, column=4 );
            cell_C4.value = v.firstname_eng;
    
            cell_C5 = sheet.cell( row=i, column=5 );
            cell_C5.value = v.lastname_eng;
    
            cell_C6 = sheet.cell( row=i, column=6 );
            cell_C6.value = v.lastname_eng;
    
            cell_C7 = sheet.cell( row=i, column=7 );
            cell_C7.value = v.sex;
            
            cell_C8 = sheet.cell( row=i, column=8 );
            cell_C8.value = v.birthdate;
    
            cell_C9 = sheet.cell( row=i, column=9 );
            cell_C9.value = v. pid;
            
            cell_C10 = sheet.cell( row=i, column=10 );
            cell_C10.value = v.passport;
            
            cell_C11 = sheet.cell( row=i, column=11 );
            cell_C11.value = v.countryname;
            
            cell_C12 = sheet.cell( row=i, column=12 );
            cell_C12.value = v.house_no;
            
            cell_C13 = sheet.cell( row=i, column=13 );
            cell_C13.value = v.building_village;
            
            cell_C14 = sheet.cell( row=i, column=14 );
            cell_C14.value = v.moo;
    
            cell_C15 = sheet.cell( row=i, column=15 );
            cell_C15.value = v.soi;
    
            cell_C16 = sheet.cell( row=i, column=16 );
            cell_C16.value = v.road;
    
            cell_C17 = sheet.cell( row=i, column=17 );
            cell_C17.value = v.county;
    
            cell_C18 = sheet.cell( row=i, column=18 );
            cell_C18.value = v.city;
    
            cell_C19 = sheet.cell( row=i, column=19 );
            cell_C19.value = v.province;
    
            cell_C20 = sheet.cell( row=i, column=20 );
            cell_C20.value = v.postcode;
    
            cell_C21 = sheet.cell( row=i, column=21 );
            cell_C21.value = v.mobile;
   
            cell_C22 = sheet.cell( row=i, column=22 );
            cell_C22.value = v.telephone;
    
            cell_C23 = sheet.cell( row=i, column=23 );
            cell_C23.value = v.email;
    
            cell_C24 = sheet.cell( row=i, column=24 );
            cell_C24.value = v.housing_type;
    
            cell_C25 = sheet.cell( row=i, column=25 );
            cell_C25.value = v.category;
    
            cell_C26 = sheet.cell( row=i, column=26 );
            cell_C26.value = v.salary;
    
            cell_C27 = sheet.cell( row=i, column=27 );
            cell_C27.value = v.education;
            
            i = i +1;
            
        
        
       
        
        wb.save(fileName);
        
    def checkEmailDuplicate(self,exportEmail):
        
        idexport = exportEmail.id_export_email;
        
        
        #check duplicate
        log.info("step one");
        self.dupEmail =  model.EmailData.checkDuplicateEmail(idexport);
        log.info("size of duplicate : " + str(len(self.dupEmail))  );
        #check not duplicate
        log.info("step two");
        self.noDupEmail =  model.EmailData.checkNotDuplicateEmail(idexport);
        log.info("size of not duplicate : " + str(len(self.dupEmail)) );
       
        #print len(self.dupEmail);
        
        #update value
        
        
       
        #update value
        log.info("step three : update export email");
        u = update(model.ExportEmail).where( model.ExportEmail.id_export_email==idexport).\
        values(same_old_row= len(self.dupEmail),insert_real_row= len(self.noDupEmail) , id_status_export= 6)

        self.session.execute(u)
        self.session.flush() ;
        self.session.commit();
        
        
        #export file
        
        if(len(self.dupEmail) > 0):
            log.info("dupEmail");
            path =exportEmail.error_path_file;
            
            if (os.path.isfile(path) ):
                workbook = openpyxl.load_workbook(filename = path,  use_iterators = False);
                worksheets = workbook.get_sheet_names();
                print worksheets;
            else:
                workbook = openpyxl.Workbook();
                sheet = workbook.get_active_sheet();
                workbook.remove_sheet(sheet);
            
            
            self.writeToExcel(workbook,self.dupEmail,path,'emailold');
       
        u = update(model.ExportEmail).where( model.ExportEmail.id_export_email==idexport).\
        values(  id_status_export= 7)

        self.session.execute(u)
        self.session.flush() ;
        self.session.commit();
        
        
        u = update(model.ExportEmail).where( model.ExportEmail.id_export_email==idexport).\
        values(  id_status_export= 8)

        self.session.execute(u)
        self.session.flush() ;
        self.session.commit();
        #insert to data temp
        for emailData in self.noDupEmail:
            emailTemp = model.EmailTemp();
            emailTemp.copyData(emailData);    
            self.session.add(emailTemp);
            self.session.flush() ;
            emailTemp = None;
        
        log.info( "insert success");    
        self.session.commit();
        log.info( "insert commit");  
        
        u = update(model.ExportEmail).where( model.ExportEmail.id_export_email==idexport).\
        values(  id_status_export= 9)

        self.session.execute(u)
        self.session.flush() ;
        self.session.commit(); 
        
        
        